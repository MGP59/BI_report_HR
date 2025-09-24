import time 
import datetime 
import pandas as pd 
from openpyxl import load_workbook 
from airflow.decorators import dag, task
from airflow.models import Variable
from sqlalchemy import create_engine 

# Параметры для аутентификации и входа

user = 'grigoriy'
host = '10.100.8.47'
db = 'internal_data'
pwd = Variable.get('planning_datas_password') 

postgresql_url = f'postgresql+psycopg2://{user}:{pwd}@{host}/{db}'
engine = create_engine(postgresql_url) 

# запрос к БД для получения различных дат увольнения сотрудников

query_db = '''SELECT t.agency, t.activ, t.department, t.group 
                FROM list_divisions_hr AS t
                GROUP BY t.agency, t.activ, t.department, t.group''' 

PATH_FILE = r'/mnt/zup_bi/Структура.xlsx' 

default_args = {
                'owner': 'airflow'
                } 

# Необходимые функции

def clean_text(x): 
    '''
    Очистка от лишних пробелов и символов
    ''' 
    if pd.notna(x): 
        x = str(x) 
        x = ' '.join(x.split()) 
        x = x.strip() 
    
    return x 

@dag(default_args=default_args,
     schedule_interval='30 9 * * *',
     start_date=datetime.datetime(2023, 1, 1),
     catchup=False, 
     )
def list_divisions_hr(): 
    @task
    def get_dis_emp_db(query: str, engine) -> pd.DataFrame: 
        """
        Функция получает таблицу с подразделениями из БД
        :param query: запрос к БД
        :return: таблица
        """
        dis_emp_db = pd.read_sql(query, engine) 
        
        return dis_emp_db 

    @task
    def get_data_divisions_hr(path_files: str, 
                            dis_emp_db: pd.DataFrame 
                            ) -> pd.DataFrame: 
        
        # Загрузка файла: 

        wb = load_workbook(filename=path_files) 
        books = [] 

        for names in wb.sheetnames: 
            sheet = wb[names] 
            for row in sheet.values: 
                books.append([value for value in row]) 
                
        # Преобразуем в датафрейм: 

        books1 = pd.DataFrame(data=books[1:], columns=['agency','activ','department','group']) 
        
        # Удалим пустые строки: 

        books1 = books1.dropna(subset=['agency']) 
        books1 = books1[books1['agency'] != 'Агентство/дирекция '].reset_index(drop=True) 
        
        # Очистим все столбцы: 

        books1 = books1.applymap(lambda x: clean_text(x)) 
        
        return books1[~((books1['agency'].isin(dis_emp_db['agency'])) \
                       & (books1['activ'].isin(dis_emp_db['activ'])) \
                       & (books1['department'].isin(dis_emp_db['department'])) \
                       & (books1['group'].isin(dis_emp_db['group'])))].reset_index(drop=True) 
    
    @task
    def append_data_to_db(data: pd.DataFrame, table: str, engine):
        """
        Функция добавляет данные из датафрейма в БД
        :param data: датафрейм
        :param table: имя таблицы
        :return: None
        """
        for i in range(len(data) // 10000 + 1):
            data.iloc[i * 10000: (i + 1) * 10000].to_sql(con=engine,
                                                         name=table,
                                                         if_exists='append',
                                                         index=False)
            time.sleep(3) 

    # Загрузка таблицы с уволенными в БД
    
    dis_emp_db = get_dis_emp_db(query=query_db, engine=engine) 
    
    df = get_data_divisions_hr(path_files=PATH_FILE, 
                                dis_emp_db=dis_emp_db) 
    
    append_data_to_db(data=df, table='list_divisions_hr', engine=engine) 

list_divisions_hr = list_divisions_hr()
