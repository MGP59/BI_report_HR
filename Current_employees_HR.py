import time 
import datetime 
import requests 
import pandas as pd 
from openpyxl import load_workbook 
from airflow.decorators import dag, task
from airflow.models import Variable
from sqlalchemy import create_engine 

# Параметры для аутентификации и входа

user = 'grigoriy'
host = '10.100.8.47'
db = 'internal_data'
pwd = Variable.get('planning_datas_password') 

postgresql_url = f'postgresql+psycopg2://{user}:{pwd}@{host}/{db}'
engine = create_engine(postgresql_url) 

PATH_FILE = r'/mnt/zup_bi/Сотрудники - актуальный список (для bi) (XLSX).xlsx' 

default_args = {'owner': 'airflow'} 

COLUMNS = ['employee', 'post', 'date_admission', 'date_birth',
            'work_experience_company_for_years', 'condition', 'agency', 'activ',
            'department', 'group1', 'group2'] 


# Необходимые функции

def clean_fio(x): 
    '''
    Очистка ФИО от лишних пробелов и символов
    ''' 
    x = x.replace('(ув.)','') 
    x = x.strip() 
    x = ' '.join([i.strip().capitalize() for i in x.split()]) 
    
    return x 

def join_name_org(x): 
    '''
    Соединим в одну строку подразделения
    '''
    x = filter(None, x) 
    x = [i for i in x if pd.notna(i)] 
    
    return '|'.join([i for i in x]) 

def datediff_years(x): 
    '''
    Число лет между двумя датами
    ''' 
    if pd.notna(x[1]) and pd.notna(x[0]): 
        n_days = (x[1]-x[0]).days 
        n_years = int(n_days / 365.25) 

        return n_years 
    return 0 

def datediff_months(x): 
    '''
    Число месяцев между двумя датами без лет
    ''' 
    if pd.notna(x[1]) and pd.notna(x[0]): 
        n_days = (x[1]-x[0]).days 
        n_years = int(n_days / 365.25) 
        n_months = n_days - n_years * 365.25 

        return int(n_months) 
    return 0 

def datediff_all_months(x): 
    '''
    Число месяцев между двумя датами 
    ''' 
    if pd.notna(x[1]) and pd.notna(x[0]): 
        n_days = (x[1]-x[0]).days 
        n_months = int(n_days / 30.4375) 

        return n_months 
    return 0 

def left_shift_list(x: list) -> list: 
    '''
    Сдвигаем значения влево с заполнением справа крайним значением
    ''' 
    lst = list(filter(None, x)) 
    additional_size = len(x) - len(lst) 
    lst.extend([lst[-1]]*additional_size) 
    
    return lst 

@dag(default_args=default_args,
     schedule_interval='30 10 * * *',
     start_date=datetime.datetime(2023, 1, 1),
     catchup=False, 
     )
def current_employees_1c(): 
    @task
    def get_data_employees(path_f: str, 
                            cols: list, 
                            engine) -> pd.DataFrame: 
        
        wb = load_workbook(filename=path_f) 
        sheet = wb[wb.sheetnames[0]] 
        books = [] 
        for row in sheet.values: 
            books.append([value for value in row]) 
            
        books1 = pd.DataFrame(books) 
        
        # Индекс первой строки, содержащей таблицу: 

        ind_row = books1[books1.isin(['Сотрудник']).any(axis=1)].index[0].min() 
        
        # Таблица с данными: 

        dis_emp = books1.loc[ind_row+1:, books1.loc[ind_row,:].notna()] 
        dis_emp.columns = books1.loc[ind_row, books1.loc[ind_row,:].notna()] 
        dis_emp = dis_emp.reset_index(drop=True) 
        dis_emp['Сотрудник'] = dis_emp['Сотрудник'].apply(lambda x: clean_fio(x)) 
        dis_emp['Дата приема'] = pd.to_datetime(dis_emp['Дата приема'], format='%d.%m.%Y') 
        
        # Переименуем столбцы: 

        dis_emp = dis_emp.rename(columns={'Сотрудник': 'employee', 
                                        'Должность': 'post', 
                                        'Дата приема': 'date_admission', 
                                        'Дата рождения': 'date_birth',
                                        'Стаж работы на предприятии лет': 'work_experience_company_for_years', 
                                        'Состояние': 'condition', 
                                        'Подразделение.Вышестоящее подразделение.Вышестоящее подразделение.Вышестоящее подразделение.Вышестоящее подразделение': 'agency',
                                        'Подразделение.Вышестоящее подразделение.Вышестоящее подразделение.Вышестоящее подразделение': 'activ',
                                        'Подразделение.Вышестоящее подразделение.Вышестоящее подразделение': 'department',
                                        'Подразделение.Вышестоящее подразделение': 'group1', 
                                        'Подразделение': 'group2'}) 
        
        # Отберем необходимые столбцы: 

        dis_emp = dis_emp[cols].reset_index(drop=True) 
        
        # Сдвинем влево и заполним право столбцы: 

        cols_val = ['agency','activ','department','group1','group2'] 
        dep_data = dis_emp[cols_val].apply(lambda x: pd.Series(left_shift_list(x)), axis=1) 
        dep_data.columns = cols_val 
        
        # Часть таблицы без подразделений: 

        dis_emp_data = dis_emp[['employee', 'post', 'date_admission', 'date_birth', 
                               'work_experience_company_for_years', 'condition']] 
        
        # Объединим в одну таблицу: 

        emp_data = pd.concat([dis_emp_data, dep_data], axis=1) 
        
        # Заполним пустые значения и приведем в необходимый формат данные: 

        emp_data['work_experience_company_for_years'] = emp_data['work_experience_company_for_years'].fillna(0) 
        emp_data['date_admission'] = pd.to_datetime(emp_data['date_admission']) 
        emp_data['date_birth'] = pd.to_datetime(emp_data['date_birth'], format='%d.%m.%Y') 
        
        emp_data = emp_data.astype({'employee': str, 
                                'post': str, 
                                'date_admission': 'datetime64[ns]', 
                                'date_birth': 'datetime64[ns]',
                                'work_experience_company_for_years': int, 
                                'condition': str, 
                                'agency': str, 
                                'activ': str,
                                'department': str, 
                                'group1': str, 
                                'group2': str}) 
        if len(emp_data) > 0: 
            
            with engine.connect() as conn: 
                delete_query = "TRUNCATE current_employees_hr" 
                conn.execute(delete_query) 
                
            return emp_data 
    
    @task
    def append_data_to_db(data: pd.DataFrame, table: str, engine):
        """
        Функция добавляет данные из датафрейма в БД
        :param data: датафрейм
        :param table: имя таблицы
        :return: None
        """
        for i in range(len(data) // 10000 + 1):
            data.iloc[i * 10000: (i + 1) * 10000].to_sql(con=engine,
                                                         name=table,
                                                         if_exists='append',
                                                         index=False)
            time.sleep(3) 

    # Загрузка таблицы актуального списка в БД
    
    df = get_data_employees(path_f=PATH_FILE, 
                            cols=COLUMNS, 
                            engine=engine) 
    
    append_data_to_db(data=df, 
                      table='current_employees_hr', 
                      engine=engine) 

current_employees_1c = current_employees_1c()
