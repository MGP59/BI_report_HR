import time 
import datetime 
import requests 
import pandas as pd 
from openpyxl import load_workbook 
from airflow.decorators import dag, task
from airflow.models import Variable
from sqlalchemy import create_engine 
from airflow.providers.telegram.hooks.telegram import TelegramHook 

# Параметры для аутентификации и входа

user = 'grigoriy'
host = '10.100.8.47'
db = 'internal_data'
pwd = Variable.get('planning_datas_password') 

postgresql_url = f'postgresql+psycopg2://{user}:{pwd}@{host}/{db}'
engine = create_engine(postgresql_url) 

# запрос к БД для получения различных дат увольнения сотрудников

query_db = '''SELECT DATE(t.date_dismissal) AS date_dismissal, t.full_name
                FROM dismissed_employees AS t
                GROUP BY t.date_dismissal, t.full_name''' 

USER_LOGIN = 'XXXXXXX\grigorij.xxxxxxx' 
USER_PASSWORD = Variable.get('deltaclick_password') 
PATH_FILE = r"/mnt/zup_bi/Уволенные сотрудники (для bi) (XLSX).xlsx" 

default_args = {
                'owner': 'airflow'
                } 

# Необходимые функции

def clean_fio(x): 
    '''
    Очистка ФИО от лишних пробелов и символов
    ''' 
    x = x.replace('(ув.)','') 
    x = x.strip() 
    x = ' '.join([i.strip().capitalize() for i in x.split()]) 
    
    return x 

def join_name_org(x): 
    '''
    Соединим в одну строку подразделения
    '''
    x = filter(None, x) 
    x = [i for i in x if pd.notna(i)] 
    
    return '|'.join([i for i in x]) 

def datediff_years(x): 
    '''
    Число лет между двумя датами
    ''' 
    if pd.notna(x[1]) and pd.notna(x[0]): 
        n_days = (x[1]-x[0]).days 
        n_years = int(n_days / 365.25) 

        return n_years 
    return 0 

def datediff_months(x): 
    '''
    Число месяцев между двумя датами без лет
    ''' 
    if pd.notna(x[1]) and pd.notna(x[0]): 
        n_days = (x[1]-x[0]).days 
        n_years = int(n_days / 365.25) 
        n_months = n_days - n_years * 365.25 

        return int(n_months) 
    return 0 

def datediff_all_months(x): 
    '''
    Число месяцев между двумя датами 
    ''' 
    if pd.notna(x[1]) and pd.notna(x[0]): 
        n_days = (x[1]-x[0]).days 
        n_months = int(n_days / 30.4375) 

        return n_months 
    return 0 

@dag(default_args=default_args,
     schedule_interval='30 12 * * *',
     start_date=datetime.datetime(2023, 1, 1),
     catchup=False, 
     )
def dismissed_employees_1c(): 
    @task
    def get_dis_emp_db(query: str, engine) -> pd.DataFrame: 
        """
        Функция получает таблицу с уволенными сотрудниками из БД
        :param query: запрос к БД
        :return: таблица с датами увольнения и ФИО
        """
        dis_emp_db = pd.read_sql(query, engine) 
        
        return dis_emp_db 

    @task
    def get_data_dismissed_employees(
                                    user_login: str, 
                                    user_password: str, 
                                    dis_emp_db: pd.DataFrame 
                                    ) -> pd.DataFrame: 
        
        wb = load_workbook(filename=PATH_FILE) 
        sheet = wb[wb.sheetnames[0]] 
        books = [] 
        for row in sheet.values: 
            books.append([value for value in row]) 

        employess2 = pd.DataFrame(books) 
        
        # employess2 = pd.read_excel('//dp1ck/zup_Bi/Уволенные сотрудники (для bi) (2023).xlsx') 
        
        # Индекс первой строки, содержащей таблицу: 

        ind_row = employess2[employess2.isin(['Организация']).any(axis=1)].index[0].min() 
        
        # Таблица с данными: 

        dis_emp = employess2.loc[ind_row+1:, employess2.loc[ind_row,:].notna()] 
        dis_emp.columns = employess2.loc[ind_row, employess2.loc[ind_row,:].notna()] 
        dis_emp = dis_emp.reset_index(drop=True) 
        dis_emp['Сотрудник'] = dis_emp['Сотрудник'].apply(lambda x: clean_fio(x)) 
        dis_emp['Дата приема'] = pd.to_datetime(dis_emp['Дата приема'], format='%d.%m.%Y') 
        dis_emp['Дата увольнения'] = pd.to_datetime(dis_emp['Дата увольнения'], format='%d.%m.%Y') 
        
        # Добавим необходимые столбцы: 

        dis_emp['agency'] = dis_emp[['Организация',
                                     'Подразделение.Вышестоящее подразделение.Вышестоящее подразделение.Вышестоящее подразделение',
                                     'Подразделение.Вышестоящее подразделение.Вышестоящее подразделение',
                                     'Подразделение.Вышестоящее подразделение']].apply(lambda x: join_name_org(x), axis=1) 
        dis_emp['start_date_work_department'] = dis_emp['Дата приема'] 
        dis_emp['work_experience_years'] = dis_emp[['Дата приема','Дата увольнения']].apply(lambda x: datediff_years(x), axis=1) 
        dis_emp['work_experience_month'] = dis_emp[['Дата приема','Дата увольнения']].apply(lambda x: datediff_months(x), axis=1) 
        dis_emp['only_months'] = dis_emp[['Дата приема','Дата увольнения']].apply(lambda x: datediff_all_months(x), axis=1) 
        dis_emp['city'] = None 
        dis_emp['comment_optional'] = None 
        
        # Переименуем столбцы: 

        dis_emp = dis_emp.rename(columns={
                                'Сотрудник': 'full_name', 
                                'Дата приема': 'date_admission', 
                                'Дата увольнения': 'date_dismissal', 
                                'Должность': 'post', 
                                'Приказ об увольнении.Причина увольнения (Увольнения).Входит в группу': 'initiator_dismissal', 
                                'Приказ об увольнении.Причина увольнения (Увольнения)': 'reason_dismissal', 
                                'Подразделение': 'department', 
                                'Организация': 'legal_entity'
                                        }) 
        
        # Отберем необходимые столбцы: 

        col = ['date_admission', 'start_date_work_department', 
                'date_dismissal', 'full_name', 'agency', 
                'department', 'work_experience_years',
                'work_experience_month', 'post',
                'only_months', 'legal_entity', 'city',
                'initiator_dismissal', 'reason_dismissal',
                'comment_optional'] 

        dis_emp = dis_emp[col] 
        
        return dis_emp[~((dis_emp['date_dismissal'].isin(dis_emp_db['date_dismissal'])) \
                  & (dis_emp['date_dismissal'].isin(dis_emp_db['date_dismissal'])))].reset_index(drop=True) 
    
    @task
    def append_data_to_db(data: pd.DataFrame, table: str, engine):
        """
        Функция добавляет данные из датафрейма в БД
        :param data: датафрейм
        :param table: имя таблицы
        :return: None
        """
        for i in range(len(data) // 10000 + 1):
            data.iloc[i * 10000: (i + 1) * 10000].to_sql(con=engine,
                                                         name=table,
                                                         if_exists='append',
                                                         index=False)
            time.sleep(3) 

    # Загрузка таблицы с уволенными в БД
    
    dis_emp_db = get_dis_emp_db(query=query_db, engine=engine) 
    df = get_data_dismissed_employees(
                                    user_login=USER_LOGIN, 
                                    user_password=USER_PASSWORD, 
                                    dis_emp_db=dis_emp_db 
                                    )
    append_data_to_db(data=df, table='dismissed_employees', engine=engine) 

dismissed_employees_1c = dismissed_employees_1c()
